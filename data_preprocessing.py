# Paqueterías
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Valida que los datos de un dataframe sean de cierto tipo especificado
def validate_dtypes(dtypes_series: pd.core.series.Series, dtype_dict: dict, date_columns: list = []) -> tuple[bool, str]:
  validation_message: str = "Successful validation"

  for column in dtypes_series.index:

    column_type = dtypes_series[column].type

    if (column in dtype_dict.keys()) and (column_type != dtype_dict[column]):
      validation_message = f"{column} type is {dtypes_series[column]}, but the expected type is {dtype_dict[column]}"
      return (False, validation_message)

    elif (column in date_columns) and (dtypes_series[column].type != np.datetime64):
      validation_message = f"{column} type is {dtypes_series[column]}, but the expected type is datetime64[ns]"
      return (False, validation_message)

  return (True, validation_message)

# Compara si los elementos de la columna de la tabla "izquierda" se encuentran
# en otra columna de la tabla "derecha"
def left_compare(left_df, left_val, right_df, right_val, indicator_name="_merge"):
  merged = pd.merge(pd.DataFrame(left_df[left_val].unique(),columns=[left_val]), right_df[right_val], left_on=left_val, right_on=right_val, how="left", indicator=indicator_name)
  merged[indicator_name] = merged[indicator_name].map({"left_only":False, "both":True})
  merged = merged.set_index(left_val)
  assert merged.shape[0] == left_df.shape[0], f"{merged.shape[0] - left_df.shape[0]}"
  return merged

# Procesamiento del reporte general de usuarios
def procesar_reporte_general_de_usuarios(path):
  objects = [
      "USER_ID",
      "ID_UNICO_ANDREA",
      "NOMBRE",
      "APELLIDO_PATERNO",
      "APELLIDO_MATERNO",
      "PERFIL",
      "CORREO_ELECTRONICO",
      "CELULAR",
      "ESTADO",
      "SEXO",
      "ESTATUS",
      "RFC"
  ]

  floats = [
      "NIVEL",
      "JOYAS_TOTALES_GANADAS",
      "JOYAS_CANJEADOS",
      "JOYAS_DISPONIBLES",
  ]

  dates = [
      "FECHA_DE_NACIMIENTO",
      "FECHA_COMPLEMENTO_DATOS",
      "ULTIMO_INGRESO_APP",
      "FECHA_CREACIÓN_DE_USUARIO"
  ]

  na_values_list = [
      "",
      "null",
      "NO UPDATE",
      "0000-00-00"
  ]

  dtype_dict = {column:np.object_ for column in objects} | {column:np.float64 for column in floats}

  RGU = pd.read_csv(
      path,
      encoding="latin-1",
      dtype=dtype_dict,
      na_values=na_values_list,
      parse_dates=dates
  )

  dtypes_validation = validate_dtypes(RGU.dtypes, dtype_dict, dates)
  assert dtypes_validation[0], dtypes_validation[1]

  RGU["Ganadoras"] = RGU.apply(lambda x: "No ganadora" if (x["JOYAS_TOTALES_GANADAS"]==0 and x["PERFIL"]!="Mayorista") else "Ganadora", axis=1)
  RGU["Con canje"] = RGU["JOYAS_CANJEADOS"].apply(lambda x: "Sin canje" if x==0 else "Con canje")
  RGU["Con ingreso"] = [np.nan] * RGU.shape[0]
  RGU.loc[pd.isnull(RGU["ULTIMO_INGRESO_APP"]), "Con ingreso"] = "Sin ingreso"
  RGU.loc[pd.notnull(RGU["ULTIMO_INGRESO_APP"]), "Con ingreso"] = "Con ingreso"

  generaciones_bins_dates = [
      "1/1/1900",
      "12/31/1964",
      "12/31/1976",
      "12/31/1995",
      "12/31/2050"
  ]

  generaciones_bins = pd.Series(generaciones_bins_dates,dtype="datetime64[ns]")

  generaciones_labels = [
      "Baby boomer",
      "Generacion X",
      "Millenial",
      "Z y otra"
  ]

  RGU["generación"] = pd.cut( RGU["FECHA_DE_NACIMIENTO"], bins=generaciones_bins, labels=generaciones_labels )

  return RGU

# Procesamiento del reporte de metas y resultados
def procesar_reporte_metas_y_resultados(path, reporte_general_de_usuarios=None):
  objects = [
      "ID_UNICO_ANDREA",
      "NOMBRE",
      "APELLIDO_PATERNO",
      "APELLIDO_MATERNO",
      "PERFIL",
      "PORCENTAJE_DE_CUMPLIMIENTO"
  ]

  floats = [
      "NIVEL",
      "HABILITADO",
      "ACTIVADO",
      "MES",
      "AÑO",
      "CUOTA_OBJETIVO",
      "MONTO_DE_VENTA_NETA_ACUMULADA_AL_CIERRE_DE_MES",
      "JOYAS_TOTALES_GANADAS_POR_COMPRAS"
  ]

  dtype_dict = {column:np.object_ for column in objects} | {column:np.float64 for column in floats}

  try:

      RMR = pd.read_csv(
          path,
          encoding="utf-8",
          dtype=dtype_dict
      )
  except UnicodeDecodeError:
      RMR = pd.read_csv(
          path,
          encoding="latin-1",
          dtype=dtype_dict
      )

  dtypes_validation = validate_dtypes(RMR.dtypes, dtype_dict)
  assert dtypes_validation[0], dtypes_validation[1]

  
  RMR["PORCENTAJE_DE_CUMPLIMIENTO"] = pd.to_numeric(RMR["PORCENTAJE_DE_CUMPLIMIENTO"].apply(lambda x: x.replace("%","") if isinstance(x,str) else x))

  RMR["Logro meta"] = RMR["PORCENTAJE_DE_CUMPLIMIENTO"].apply(lambda x: "Cumplió" if x >= 100 else "No cumplió")

  RMR["Crecimiento sobre la renta"] = RMR["MONTO_DE_VENTA_NETA_ACUMULADA_AL_CIERRE_DE_MES"] - RMR["CUOTA_OBJETIVO"]

  #return pd.merge(RMR, left_compare(RMR, "ID_UNICO_ANDREA", reporte_general_de_usuarios, "ID_UNICO_ANDREA", indicator_name="Participa"), on="ID_UNICO_ANDREA")
  return RMR 

# Procesamiento de datos de Shipping List
def procesar_shipping_list(path):
  objects = [
      "ID_UNICO_CANJE",
      "PEDIDO EREWARD",
      "EREWARD CANJE ID",
      "NOMBRE (S)",
      "APELLIDO PATERNO",
      "APELLIDO MATERNO",
      "ID_UNICO_ANDREA",
      "PERFIL",
      "CELULAR CANJE",
      "CORREO CANJE",
      "CORREO REGISTRADO",
      "CATEGORIA",
      "CLASIFICACION DE PRODUCTO",
      "NOMBRE CORTO",
      "SKU",
      "CODIGO PMR",
      "MARCA",
      "TELEFONO_DE_RECARGA",
      "COMPAÑÍA RECARGA",
      "PAQUETERIA",
      "LIGA_POD",
      "CALLE",
      "REFERENCIAS",
      "COLONIA",
      "MUNICIPIO",
      "ESTADO",
      "RFC MONEDERO",
      "NOMBRE COMPLETO MONEDERO",
      "LUGAR DE CANJE",
      "GUIA",
      "ESTATUS DE ENTREGA",
      "ID_ARTICULO_SAP",
      "QUIEN RECIBIO",
      "NÚMERO EXTERIOR",
      "NÚMERO INTERIOR"
  ]

  floats = [
      "PEDIDO",
      "PUNTOS",
      "CANTIDAD",
      "CODIGO POSTAL",
      "PRECIO LOGISTICA",
      "FEE",
      "PRECIO VENTA INTEGRADO",
      "PRECIO DE VENTA MAS AJUSTE",
      "MONTO DE DISPERSION",
      "COMISION PROVEEDOR",
      "PRECIO PLASTICO"
  ]

  dates = [
      "FECHA_DE_CANJE",
      "FECHA DE CORTE",
      "FECHA ESTIMADA DE ENTREGA",
      "FECHA DE RECEPCIÓN"
  ]

  na_values_list = ["", "NA", "s/n", np.nan, np.float64("nan")]

  dtype_dict = {column:np.object_ for column in objects} | {column:np.float64 for column in floats}

  try:
    SL = pd.read_csv(
        path,
        encoding="latin-1",
        dtype=dtype_dict,
        na_values=na_values_list,
    )
    SL[dates] = SL[dates].astype("datetime64[ns]")
  except KeyError:
    SL = pd.read_csv(
        path,
        encoding="utf-8",
        dtype=dtype_dict,
        na_values=na_values_list,
    )
    SL[dates] = SL[dates].astype("datetime64[ns]")
      
  SL["PRECIO PRODUCTO"] = SL["PRECIO PRODUCTO"].apply(lambda x: x.replace("$", "")).astype("float64")

  dtypes_validation = validate_dtypes(SL.dtypes, dtype_dict, dates)
  assert dtypes_validation[0], dtypes_validation[1]
  return SL

def conteo_distintivo(dataframe, column, count_name="COUNT",guardar_como=None):
  print(column,end="\n\n")
  distinct_count_df = pd.DataFrame([dataframe[column].nunique()], columns=[count_name])
  if guardar_como is not None:
    distinct_count_df.to_csv(guardar_como, encoding="latin-1", index=False)
  return distinct_count_df

def porcentaje_valores_dist(dataframe, column, decimals=2, plot_percentages=False, plot_type=None, guardar_como=None):
  count_df = round(dataframe[column].value_counts(1)*100, decimals)

  if plot_percentages:
    if plot_type == None:
      plot_type = "bar"

    if plot_type == "bar":
      ax0 = count_df.plot.bar(rot=False)
      for p0 in ax0.patches:
        ax0.annotate(str(round(p0.get_height(),2))+"%", (p0.get_x() + p0.get_width() / 2., p0.get_height()), ha='center', va='bottom')
      plt.show()

    elif plot_type == "pie":
      wedges, texts, autotexts = plt.pie( x=count_df, shadow=False, autopct='%1.0f%%', pctdistance=1.1 )
      plt.legend(wedges, count_df.index,
                title="Porcentajes",
                loc="center left",
                bbox_to_anchor=(1, 0, 0.5, 1))
      plt.show()
    else:
      print("Tipo de gráfico no reconocido o on disponible. Solo se acepta plot_type = 'bar' o 'pie'.")

  count_df = count_df.rename("Porcentaje %")
  if guardar_como is not None:
      count_df.to_csv(guardar_como, encoding="latin-1", index=True)
  return count_df

def mostrar_tabla(dataframe):
  with pd.option_context('display.max_rows', None, 'display.max_columns', None):
    print(dataframe)

def tabla_pivote(dataframe, filas, valores=None, columnas=None, margins=True, margins_name="Total", aggfunc=None, rename_cols=None, return_=False, guardar_como=None):
  if valores is None:
    if columnas is None:
      frame_to_return = dataframe[valores].describe()
      if guardar_como is not None: frame_to_return.to_csv(guardar_como, encoding="latin-1", index=True)
      return frame_to_return
    else:
      if (dataframe[columnas].dtypes.name == "object") or (dataframe[columnas].dtypes.name == "datetime64[ns]"):
        dataframe_copy = dataframe[[valores,columnas]].copy()
        dataframe_copy["aux_column"] = range( dataframe_copy.shape[0] )
        pivot_table = pd.pivot_table(
            dataframe_copy,
            index=filas,
            values="aux_column",
            columns=columnas,
            aggfunc="nunique",
            margins=margins,
            margins_name=margins_name
        )
        if guardar_como is not None:
            pivot_table.to_csv(guardar_como, encoding="latin-1", index=True)
        return pivot_table
        
      else:
        raise TypeError("Las columnas deben de contener solo valores categóricos, no numéricos.")

  if not aggfunc:
      if dataframe[valores].dtypes.name == "object":
          aggfunc = "nunique"
      elif (dataframe[valores].dtypes.name == "float64") or (dataframe[valores].dtypes.name == "int64"):
          aggfunc = "sum"
      elif isinstance(valores, list):
        aggfunc = {}
        for column in valores:
            if (dataframe[column].dtypes.name == "object") or (dataframe[column].dtypes.name == "datetime64[ns]"): aggfunc[column] = "nunique"
            else: aggfunc[column] = "sum"
                
  elif isinstance(aggfunc,dict):
      for valor in valores:
          if valor not in aggfunc.keys():
              if (dataframe[valor].dtypes.name == "object") or (dataframe[valor].dtypes.name == "datetime64[ns]"): aggfunc[valor] = "nunique"
              else: aggfunc[valor] = "sum"

  pivot_table = pd.pivot_table(
        dataframe,
        index=filas,
        values=valores,
        columns=columnas,
        aggfunc=aggfunc,
        margins=margins,
        margins_name=margins_name
    )

  if rename_cols:
    pivot_table = pivot_table.rename(columns=rename_cols)

  if guardar_como is not None:
      pivot_table.to_csv(guardar_como, encoding="latin-1", index=True)
  return pivot_table

def filtrar_Y(dataframe, *condiciones, guardar_como=None):
  filtered_dataframe = dataframe
  FILTER = [True] * dataframe.shape[0]
  for condicion in condiciones:
    assert isinstance(condicion, tuple), 'Las condiciones deben de estar escritas de la forma:\n(columna, condicion, valor)'
    if len(condicion) == 3:
        column, cond, val = condicion
        if cond == "=":
          FILTER &= ( filtered_dataframe[column]==val ).values
    
        elif cond == ">":
          FILTER &= ( filtered_dataframe[column]>val ).values
    
        elif cond == ">=":
          FILTER &= ( filtered_dataframe[column]>=val ).values
    
        elif cond == "<":
          FILTER &= ( filtered_dataframe[column]<val ).values
    
        elif cond == "<=":
          FILTER &= ( filtered_dataframe[column]<=val ).values
    
        elif cond == "<>":
          FILTER &= ( filtered_dataframe[column]!=val ).values

        else:
            print(f"No se reconoce la condición {cond}.")
            print(f'La segunda entrada de la condición {condicion} debe de ser: "="(igual), ">"(mayor), ">="(mayor o igual), "<"(menor), "<="(menor o igual) o "<>"(diferente)')
    elif len(condicion)==2:
        if condicion[1] == "notnull":
          FILTER &= ( pd.notnull(filtered_dataframe[ condicion[0] ]) ).values
        elif condicion[1] == "isnull":
          FILTER &= pd.isnull(filtered_dataframe[ condicion[0] ])
        else:
          print(f'La segunda entrada de la condición {condicion} debe de ser: "notnull" o isnull.')

  if guardar_como is not None: filtered_dataframe[FILTER].to_csv(guardar_como, encoding="latin-1", index=False)
  return filtered_dataframe[FILTER]

def filtrar_O(dataframe, *condiciones, guardar_como=None):
  filtered_dataframe = dataframe
  FILTER = [False] * dataframe.shape[0]
  for condicion in condiciones:
    assert isinstance(condicion, tuple), 'Las condiciones deben de estar escritas de la forma:\n(columna, condicion, valor) o (columna, "notnull")'
    if len(condicion) == 3:
        column, cond, val = condicion
        if cond == "=":
          FILTER |= ( filtered_dataframe[column]==val ).values
    
        elif cond == ">":
          FILTER |= ( filtered_dataframe[column]>val ).values
    
        elif cond == ">=":
          FILTER |= ( filtered_dataframe[column]>=val ).values
    
        elif cond == "<":
          FILTER |= ( filtered_dataframe[column]<val ).values
    
        elif cond == "<=":
          FILTER |= ( filtered_dataframe[column]<=val ).values
    
        elif cond == "<>":
          FILTER |= ( filtered_dataframe[column]!=val ).values
            
        else:
            print(f"No se reconoce la condición {cond}.")
            print(f'La segunda entrada de la condición {condicion} debe de ser: "="(igual), ">"(mayor), ">="(mayor o igual), "<"(menor), "<="(menor o igual), "<>"(diferente).')
    elif len(condicion) == 2:
        if condicion[1] == "notnull":
          FILTER |= ( pd.notnull(filtered_dataframe[ condicion[0] ]) ).values
        elif condicion[1] == "isnull":
          FILTER |= pd.isnull(filtered_dataframe[ condicion[0] ])
        else:
          print(f'La segunda entrada de la condición {condicion} debe de ser: "notnull" o isnull.')

  if guardar_como is not None: filtered_dataframe[FILTER].to_csv(guardar_como, encoding="latin-1", index=False)
  return filtered_dataframe[FILTER]

def filtrar_cruzado(dataframe_1, column_1, dataframe_2, column_2=None, guardar_como=None):
  if column_2 is None: column_2 = column_1
  dataframe_to_return = dataframe_1[dataframe_1[column_1].isin( dataframe_2[column_2] )]
  if guardar_como is not None: dataframe_to_return.to_csv( guardar_como, encoding="latin-1", index=True )
  return dataframe_to_return

def suma(dataframe, *columnas, guardar_como=None):
  dataframe_to_return = pd.DataFrame({f"suma total de {columna}":[dataframe[columna].sum()] for columna in columnas})
  if guardar_como is not None: dataframe_to_return.to_csv( guardar_como, encoding="latin-1", index=False )
  return dataframe_to_return

def procesar_datos(reporte_general_de_usuarios, reporte_de_metas_y_resultados, reporte_SL, filtrar_default=True, guardar=False):
    RGU = procesar_reporte_general_de_usuarios(reporte_general_de_usuarios)
    RMR = procesar_reporte_metas_y_resultados(reporte_de_metas_y_resultados, RGU)
    SL = procesar_shipping_list(reporte_SL)
    

    print("\n¡PROCESAMIENTO DE DATOS EXITOSO!")

    if filtrar_default:
        RGU = filtrar_O(RGU, ("PERFIL","=","Estrella"), ("PERFIL","=","Mayorista"))
        RMR = filtrar_O(RMR, ("PERFIL","=","Estrella"), ("PERFIL","=","Mayorista"))
        SL = filtrar_cruzado(SL, "ID_UNICO_ANDREA", RGU, "ID_UNICO_ANDREA")

    if guardar:
        RGU.to_csv("Reportegeneraldeusuarios_procesado.csv",encoding="latin-1", index=False)
        RMR.to_csv("Reportedemetasyresultados_procesado.csv", encoding="latin-1", index=False)
        SL.to_csv("v_sl_procesado.csv", encoding="latin-1", index=False)

    return RGU, RMR, SL

def cruzar( dataframe1, dataframe2, col_tabla_izquierda, col_tabla_derecha, sufijos=None, metodo_cruce="ambas" ):
    how_options =  ["ambas","izquierda","derecha","izq-der"]
    how_dict = {"ambas":"inner","izquierda":"left","derecha":"right","izq-der":"outer"}
    assert (metodo_cruce in how_options), f"El parámetro 'metodo_cruce' debe coincidir con alguna de las siguientes opciones: {how_options}."
        
    if sufijos is not None:
        return dataframe1.merge( dataframe2, left_on=col_tabla_izquierda, right_on=col_tabla_derecha, suffixes=sufijos, how=how_dict[metodo_cruce] )
    else:
        return dataframe1.merge( dataframe2, left_on=col_tabla_izquierda, right_on=col_tabla_derecha, how=how_dict[metodo_cruce] )

def seleccionar_list(dataframe, lista_columnas):
    return dataframe[lista_columnas]

def seleccionar(dataframe, *columnas):
    return dataframe[list(columnas)]

def concatenar( *dataframes_list ):
    return pd.concat( list(dataframes_list) )

def guardar_tabla(tabla: pd.DataFrame, nombre_tabla: str, guardar_en=None) -> None:
    whole_name = nombre_tabla if nombre_tabla.endswith(".csv") else nombre_tabla + ".csv"
    if guardar_en is not None:
        whole_name = guardar_en + whole_name
    tabla.to_csv(whole_name, encoding="utf-8", index=False)

## Top usuarias

def top_usuarias(every_csv_file):
    reports: list[pd.DataFrame] = []
    columns_to_keep: list[str] = [
    "MONTO_DE_VENTA_NETA_ACUMULADA_AL_CIERRE_DE_MES",
    "NIVEL",
    "PERFIL",
    "CUOTA_OBJETIVO",
    "ID_UNICO_ANDREA",
    "MES"
    ]
    
    months = []
    for file_ in every_csv_file:
        encoding_method = "utf-8"
        try:
            report = pd.read_csv(file_, encoding=encoding_method)
        except UnicodeDecodeError:
            encoding_method = "latin-1"
            report = pd.read_csv(file_, encoding=encoding_method)
        month = str( int( report["MES"].unique()[0] ) )
        months.append(month)
        rename_columns: dict[str:str] = {"MONTO_DE_VENTA_NETA_ACUMULADA_AL_CIERRE_DE_MES":f"VENTA_NETA_MES_{month}"}
    
        report_copy = report.loc[report["PERFIL"]=="Estrella", columns_to_keep].rename(columns=rename_columns)
        reports.append(report_copy)
    
    raw_selling_details: pd.DataFrame = pd.concat( reports ).fillna(0).drop(["PERFIL","MES"], axis=1)
    
    users_selling_details = raw_selling_details.groupby( "ID_UNICO_ANDREA" ).agg( {f"VENTA_NETA_MES_{i+1}":"sum" for i in range(3)}|{"CUOTA_OBJETIVO":"mean", "NIVEL":"mean"} )
    for month in months:
        users_selling_details.loc[:,f"SUPERO_CUOTA_OBJETIVO_MES_{month}_check"] = users_selling_details.apply( lambda x: 1 if x[f"VENTA_NETA_MES_{month}"]>= x["CUOTA_OBJETIVO"] else 0, axis=1 )
    
    users_selling_details.loc[:,"VENTA_TOT"] = users_selling_details[[f"VENTA_NETA_MES_{month}" for month in months]].sum(axis=1)
    
    users_selling_details.loc[:,"TOT_PARTICIPATION"] = users_selling_details[ [f"SUPERO_CUOTA_OBJETIVO_MES_{month}_check" for month in months] ].sum( axis=1 )
    
    users_selling_details.loc[:,"(Promedio) PORCENTAJE_DE_CUMPLIMIENTO"] = 100 * users_selling_details["VENTA_TOT"] / (3*users_selling_details["CUOTA_OBJETIVO"])
    
    wb = Workbook()
    
    levels = users_selling_details["NIVEL"].unique()
    for nivel in sorted(levels):
        ws_level = wb.create_sheet(f"TOP_NIVEL_{int(nivel)}")
        top_users = users_selling_details[ (users_selling_details["NIVEL"]==nivel) & (users_selling_details["TOT_PARTICIPATION"]>0) & (users_selling_details["VENTA_TOT"]>0) ].sort_values(by=["TOT_PARTICIPATION","(Promedio) PORCENTAJE_DE_CUMPLIMIENTO"], ascending=False).reset_index().head(10)
        for row_idx, row in enumerate( dataframe_to_rows( top_users, header=True, index=False ) ):
            for column_idx, value in enumerate(row):
                ws_level.cell(row = row_idx + 1, column=column_idx + 1, value=value)     

    wb.remove(wb['Sheet'])
    wb.save(f'Top_users_months_{"_".join(months)}.xlsx')
    
    end = time()
    duration = end-start
    print(F"\nElapsed Time {duration//60} min {duration%60:0.2f} sec")

    return None